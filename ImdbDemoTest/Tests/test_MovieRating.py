from Utilities.BaseClass import BaseClass
from PageObjects.Home import Home
import openpyxl

class Test_One(BaseClass):

    def test_getMovieRating(self):
        obj= Home(self.driver)
        movieName="Ra.One"
        rating = obj.movieRating(movieName)
        print("The IMDB rating of "+movieName+" is : "+rating)

    def test_getAllMovieRating(self):
        excel = openpyxl.load_workbook("C:\\Users\\pratiyuk\\Pictures\\MovieList.xlsx")
        sheet = excel.active
        list=[]
        i=1
        while i > 0:
            if sheet.cell(row=i+1, column=1).value is not None:
                list.append(sheet.cell(row=i+1, column=1).value)
                i=i+1
            else:
                break
        obj = Home(self.driver)
        ratings = obj.allMovieRating(list)
        print(ratings)
        print(list)

        for j in range(1,len(ratings)+1):
            sheet.cell(row=j+1, column=3).value=ratings[j-1]
        excel.save("C:\\Users\\pratiyuk\\Pictures\\MovieList.xlsx")





