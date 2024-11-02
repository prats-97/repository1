import time

import openpyxl
import pytest

from POM.Home import Home
from Utilities.BaseClass import Baseclass


class Test_One(Baseclass):

    def test_getPouplarAlbum(self):
        homeObj = Home(self.driver)
        list = homeObj.getAlbumsList()

        print("List of Popular Albums on Spotify currently : ")
        for album in list:
            name = album.text
            if name != "":
                print(name)

    def test_AllPopularAlbum(self):
        homeObj = Home(self.driver)
        seeAlllist = homeObj.getSeeAllList()
        print("Complete list of Popular Albums on Spotify currently : ")
        for album in seeAlllist:
            name = album.text
            if name != "":
                print(name)
        # Below steps are saving the complete list in an Excel
        myExcel = openpyxl.load_workbook("C:\\Users\\pratiyuk\\Pictures\\PopularAlbumList.xlsx")
        sheet = myExcel.active
        for i in range(1, len(seeAlllist) + 1):
            if seeAlllist[i - 1].text != "":
                sheet.cell(row=i + 1, column=1).value = i
                sheet.cell(row=i + 1, column=2).value = seeAlllist[i - 1].text
        myExcel.save("C:\\Users\\pratiyuk\\Pictures\\PopularAlbumList.xlsx")

#    def test_storeAlbumExcel(self):
 #       myExcel = openpyxl.load_workbook("C:\\Users\\pratiyuk\\Pictures\\PopularAlbumList.xlsx")
  #      sheet = myExcel.active
   #     for i in range(1, len(self.fullList) + 1):
    #        if self.fullList[i-1].text != "":
     #           sheet.cell(row=i+1, column=1).value = i
      #          sheet.cell(row=i+1, column=2).value = self.fullList[i - 1].text

       # myExcel.save("C:\\Users\\pratiyuk\\Pictures\\PopularAlbumList.xlsx")
